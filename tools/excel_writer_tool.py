"""
Excel writer — turn the aggregated extraction results into a polished
``.xlsx`` workbook.

The writer:

- Uses ``pandas`` to build the dataframe.
- Uses ``openpyxl`` (via pandas' engine) for header styling, frozen
  panes, alternating row colours and auto-sized columns.
- Writes to ``outputs/skillflow_results_<timestamp>.xlsx``.
- Returns ``{"excel_path": ..., "rows_written": ...}`` so the caller
  can render a download button or log the path.
- Never raises on bad input — returns ``rows_written=0`` instead.
"""

from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from config import settings
from tools import register_tool


# Ordered list of columns to render in the Excel sheet.
EXCEL_COLUMNS: List[str] = [
    "Query",
    "Website",
    "URL",
    "Title",
    "Price",
    "Rating",
    "Availability",
    "Location",
    "Specs",
    "Confidence Score",
    "Fallback Used",
    "Status",
    "Method",
    "Timestamp",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _safe_filename(query: str, ts: str) -> str:
    """Build a filesystem-safe, reasonably unique workbook name."""
    slug = re.sub(r"[^A-Za-z0-9]+", "_", (query or "").strip())[:40].strip("_").lower()
    if slug:
        return f"skillflow_results_{slug}_{ts}.xlsx"
    return f"skillflow_results_{ts}.xlsx"


def _coerce_rows(rows: List[Any], query: str) -> List[Dict[str, Any]]:
    """Normalise input rows to the EXCEL_COLUMNS schema."""
    out: List[Dict[str, Any]] = []
    for row in rows or []:
        # If it's a Pydantic ExtractedResult, project via to_excel_row.
        if hasattr(row, "to_excel_row"):
            try:
                out.append(row.to_excel_row())
                continue
            except Exception:
                pass

        # If it's already a dict in flat form (e.g. result_aggregator output),
        # map known keys onto our column names.
        if isinstance(row, dict):
            specs_value = row.get("specs") or row.get("specs_raw")
            if isinstance(specs_value, dict):
                specs_text = ", ".join(f"{k}: {v}" for k, v in specs_value.items())
            else:
                specs_text = specs_value or ""

            out.append(
                {
                    "Query":            query,
                    "Website":          row.get("domain") or row.get("website", ""),
                    "URL":              row.get("url", ""),
                    "Title":            row.get("title") or "",
                    "Price":            row.get("price") or "",
                    "Rating":           row.get("rating") or "",
                    "Availability":     row.get("availability") or "",
                    "Location":         row.get("location") or "",
                    "Specs":            specs_text,
                    "Confidence Score": round(float(row.get("confidence_score") or 0.0), 2),
                    "Fallback Used":    "Yes" if row.get("fallback_used") else "No",
                    "Status":           row.get("status", ""),
                    "Method":           row.get("method") or row.get("page_method", ""),
                    "Timestamp":        row.get("timestamp")
                                          or datetime.utcnow().isoformat(timespec="seconds") + "Z",
                }
            )

    return out


def _apply_styling(workbook, sheet, n_rows: int, n_cols: int) -> None:
    """Apply header styling, alt-row colours, freeze panes and column widths."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError:
        return  # openpyxl missing for some odd reason — skip styling.

    # ------- Header --------
    header_fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="FFB7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, n_cols + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # ------- Body: alternating row colours -------
    band_fill = PatternFill(start_color="FFF2F7FB", end_color="FFF2F7FB", fill_type="solid")
    for row_idx in range(2, n_rows + 2):  # +2 because openpyxl is 1-indexed and we have a header
        for col_idx in range(1, n_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = band_fill

    # ------- Freeze top row -------
    sheet.freeze_panes = "A2"

    # ------- Auto column widths (capped) -------
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, n_rows + 2):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            length = len(str(val))
            if length > max_len:
                max_len = length
        # Clamp between 12 and 60 characters.
        sheet.column_dimensions[letter].width = max(12, min(max_len + 2, 60))

    # Sensible row height for header.
    sheet.row_dimensions[1].height = 28


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
@register_tool(
    "excel_writer",
    description="Write aggregated workflow rows to a styled Excel workbook.",
    supported_tasks=["generate_report", "export_excel"],
    input_schema={"rows": "list", "query": "str", "output_dir": "str", "filename": "str"},
    output_schema={"excel_path": "str", "rows_written": "int", "error": "str"},
)
def write_excel(
    rows: List[Any],
    *,
    query: str = "",
    output_dir: Optional[str] = None,
    filename: Optional[str] = None,
) -> Dict[str, Any]:
    """Write rows to a styled .xlsx and return ``{"excel_path", "rows_written"}``.

    On any failure the function returns ``rows_written=0`` and an
    ``error`` message instead of raising.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = output_dir or str(settings.OUTPUTS_DIR)
    os.makedirs(out_dir, exist_ok=True)
    file_name = filename or _safe_filename(query, timestamp)
    excel_path = os.path.join(out_dir, file_name)

    coerced = _coerce_rows(rows or [], query)

    if not coerced:
        return {
            "excel_path": "",
            "rows_written": 0,
            "error": "No rows to write.",
        }

    try:
        import pandas as pd  # type: ignore
    except ImportError:
        return {"excel_path": "", "rows_written": 0, "error": "pandas not installed."}

    try:
        df = pd.DataFrame(coerced, columns=EXCEL_COLUMNS)
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Results", index=False)
            workbook = writer.book
            sheet = writer.sheets["Results"]
            _apply_styling(workbook, sheet, n_rows=len(df), n_cols=len(EXCEL_COLUMNS))
    except Exception as exc:  # pragma: no cover - defensive
        return {"excel_path": "", "rows_written": 0, "error": f"Excel write failed: {exc}"}

    return {
        "excel_path": excel_path,
        "rows_written": len(coerced),
        "error": "",
    }
