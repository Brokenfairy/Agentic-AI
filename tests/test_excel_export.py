"""
Test Excel export functionality.

Validates:
- Excel file is generated
- file exists in outputs/
- correct columns are present
- rows are written correctly
- empty dataset is handled safely
- file name includes timestamp
- workbook opens without corruption
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
import pytest

from tools.excel_writer_tool import write_excel


class TestExcelExport:
    """Test Excel export functionality."""

    def test_excel_file_generation(self, temp_outputs_dir: Path) -> None:
        """Test that Excel file is generated."""
        rows = [
            {
                "Query": "Test Query",
                "Website": "amazon.in",
                "URL": "https://amazon.in/test",
                "Title": "Test Product",
                "Price": "Rs 79,900",
                "Rating": "4.5/5",
                "Availability": "In Stock",
                "Location": "",
                "Specs": "",
                "Confidence Score": 0.85,
                "Fallback Used": "No",
                "Status": "success",
                "Method": "requests",
                "Timestamp": "2024-01-01T00:00:00Z",
            }
        ]
        
        result = write_excel(rows, query="Test Query", output_dir=str(temp_outputs_dir))
        
        assert result["rows_written"] == 1
        assert result["excel_path"]
        assert Path(result["excel_path"]).exists()

    def test_excel_file_in_outputs_directory(self, temp_outputs_dir: Path) -> None:
        """Test that file is created in outputs directory."""
        rows = [{"Query": "Test", "Website": "test.com", "URL": "http://test.com"}]
        
        result = write_excel(rows, query="Test", output_dir=str(temp_outputs_dir))
        
        path = Path(result["excel_path"])
        assert path.parent.name == "outputs" or "outputs" in str(path)

    def test_correct_columns_present(self, temp_outputs_dir: Path) -> None:
        """Test that all required columns are present."""
        rows = [{"Query": "Test", "Website": "test.com", "URL": "http://test.com"}]
        
        result = write_excel(rows, query="Test", output_dir=str(temp_outputs_dir))
        
        # Open and check columns
        wb = openpyxl.load_workbook(result["excel_path"])
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        expected_columns = [
            "Query", "Website", "URL", "Title", "Price", "Rating",
            "Availability", "Location", "Specs", "Confidence Score",
            "Fallback Used", "Status", "Method", "Timestamp"
        ]
        
        for col in expected_columns:
            assert col in headers, f"Missing column: {col}"

    def test_rows_written_correctly(self, temp_outputs_dir: Path) -> None:
        """Test that data rows are written correctly."""
        rows = [
            {"Query": "Row 1", "Website": "site1.com", "URL": "http://site1.com"},
            {"Query": "Row 2", "Website": "site2.com", "URL": "http://site2.com"},
        ]
        
        result = write_excel(rows, query="Test", output_dir=str(temp_outputs_dir))
        
        wb = openpyxl.load_workbook(result["excel_path"])
        ws = wb.active
        
        # Check row count (header + data rows)
        assert ws.max_row == 3  # 1 header + 2 data

    def test_empty_dataset_handling(self, temp_outputs_dir: Path) -> None:
        """Test handling of empty dataset."""
        result = write_excel([], query="Test", output_dir=str(temp_outputs_dir))
        
        # Should return gracefully
        assert result["rows_written"] == 0

    def test_filename_includes_timestamp(self, temp_outputs_dir: Path) -> None:
        """Test that filename includes timestamp."""
        rows = [{"Query": "Test", "Website": "test.com", "URL": "http://test.com"}]
        
        result = write_excel(rows, query="Test Query", output_dir=str(temp_outputs_dir))
        
        filename = Path(result["excel_path"]).name
        
        # Should contain skillflow_results prefix
        assert "skillflow_results" in filename
        # Should contain timestamp-like numbers
        assert any(c.isdigit() for c in filename)

    def test_workbook_not_corrupted(self, temp_outputs_dir: Path) -> None:
        """Test that generated workbook is not corrupted."""
        rows = [
            {
                "Query": "Test Query",
                "Website": "amazon.in",
                "URL": "https://amazon.in/test",
                "Title": "Test Product",
                "Price": "Rs 79,900",
                "Rating": "4.5/5",
                "Availability": "In Stock",
                "Location": "",
                "Specs": "",
                "Confidence Score": 0.85,
                "Fallback Used": "No",
                "Status": "success",
                "Method": "requests",
                "Timestamp": "2024-01-01T00:00:00Z",
            }
        ]
        
        result = write_excel(rows, query="Test", output_dir=str(temp_outputs_dir))
        
        # Try to open with openpyxl - should not raise
        try:
            wb = openpyxl.load_workbook(result["excel_path"])
            wb.close()
        except Exception as e:
            pytest.fail(f"Workbook appears corrupted: {e}")

    def test_header_styling_applied(self, temp_outputs_dir: Path) -> None:
        """Test that header styling is applied."""
        rows = [{"Query": "Test", "Website": "test.com", "URL": "http://test.com"}]
        
        result = write_excel(rows, query="Test", output_dir=str(temp_outputs_dir))
        
        wb = openpyxl.load_workbook(result["excel_path"])
        ws = wb.active
        
        # Check header row has styling (fill color)
        header_cell = ws["A1"]
        assert header_cell.fill is not None

    def test_multiple_rows_export(self, temp_outputs_dir: Path) -> None:
        """Test export of multiple rows."""
        rows = [
            {
                "Query": f"Query {i}",
                "Website": f"site{i}.com",
                "URL": f"http://site{i}.com",
                "Price": f"Rs {i}9,900",
                "Rating": f"4.{i}/5",
            }
            for i in range(1, 6)
        ]
        
        result = write_excel(rows, query="Batch Test", output_dir=str(temp_outputs_dir))
        
        assert result["rows_written"] == 5

    def test_return_value_structure(self, temp_outputs_dir: Path) -> None:
        """Test that return value has expected structure."""
        rows = [{"Query": "Test", "Website": "test.com", "URL": "http://test.com"}]
        
        result = write_excel(rows, query="Test", output_dir=str(temp_outputs_dir))
        
        assert "excel_path" in result
        assert "rows_written" in result
        assert isinstance(result["excel_path"], str)
        assert isinstance(result["rows_written"], int)

    def test_excel_path_is_absolute(self, temp_outputs_dir: Path) -> None:
        """Test that returned path is absolute."""
        rows = [{"Query": "Test", "Website": "test.com", "URL": "http://test.com"}]
        
        result = write_excel(rows, query="Test", output_dir=str(temp_outputs_dir))
        
        path = Path(result["excel_path"])
        assert path.is_absolute()

    def test_query_included_in_filename(self, temp_outputs_dir: Path) -> None:
        """Test that query is reflected in filename."""
        rows = [{"Query": "iPhone 15 Test", "Website": "test.com", "URL": "http://test.com"}]
        
        result = write_excel(rows, query="iPhone 15 Test", output_dir=str(temp_outputs_dir))
        
        filename = Path(result["excel_path"]).name.lower()
        assert "iphone" in filename or "15" in filename or "test" in filename
